from openpyxl import Workbook, load_workbook
import time
import os
import xlwings

def main():

    usuario = str(input("Qual o seu usuário?\n"))
    n_Obra = str(input("\nQual o número da obra?\n"))
    
    #Base de dados
    base = load_workbook("Base de Dados - Obras.xlsx")
    PDT = base["PARAFUSOS DT"]
    PRC = base["PARAFUSOS RC"]
    
    caminho_saida =open("C:\\Users\\{}\\Documents\\Medição\\{}\\Saida.txt".format(usuario, n_Obra), 'w') # Saída do resultado
    caminho_entradaDT = open("C:\\Users\\{}\\Documents\\Medição\\{}\\EntradaDT.txt".format(usuario, n_Obra), 'r') #Entrada de informações
    caminho_entradaRC = open("C:\\Users\\{}\\Documents\\Medição\\{}\\EntradaRC.txt".format(usuario, n_Obra), 'r') #Entrada de informações
    conteudo_arquivoDT = caminho_entradaDT.read()
    conteudo_arquivoRC = caminho_entradaRC.read()
    splitDT = conteudo_arquivoDT.split("\n\n\n")
    splitRC = conteudo_arquivoRC.split("\n\n\n")

    def DT():
        #Dividir níveis de Estrutura
        First_lvl = splitDT[0]
        Second_lvl = splitDT[1]

        #Dividir MT e BT de cada Nível
        MT_BT_lvl1 = First_lvl.split("\n\n")
        MT_BT_lvl2 = Second_lvl.split("\n\n")
        
        #MT e BT de cada Nível
        Estruturas_MT_lvl1 = MT_BT_lvl1[0].split("\n")
        Estruturas_BT_lvl1 = MT_BT_lvl1[1].split("\n")
        Equipamentos = MT_BT_lvl1[2].split("\n")
        Estruturas_MT_lvl2 = MT_BT_lvl2[0].split("\n")
        Estruturas_BT_lvl2 = MT_BT_lvl2[1].split("\n")

        #Número de estruturas
        n_estruturas_MT_lvl1 = int(len(Estruturas_MT_lvl1)-2)
        n_estruturas_BT_lvl1 = int(len(Estruturas_BT_lvl1)-1)
        n_equipamentos = int(len(Equipamentos)-1)
        n_estruturas_MT_lvl2 = int(len(Estruturas_MT_lvl2)-2)
        n_estruturas_BT_lvl2 = int(len(Estruturas_BT_lvl2)-1)

        #Estruturas MT primária
        for i in range (n_estruturas_MT_lvl1):
            classificação = Estruturas_MT_lvl1[i+2].split(" - ")
            Estrutura = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])

            #Escrever na Base
            for celulal in PDT['A']:
                linha = celulal.row
                Tipo = PDT['A{}'.format(linha)].value
                if linha <= 2:
                    time.sleep(0.0000001)
                elif linha == 65:
                    break
                else:
                    if Tipo == "ESTRUTURAS E OS POSTES ONDE ESTÃO INSTALADOS DO 1º NÍVEL" or Tipo == "ESTRUTURA DE MT":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PDT['B2'].value:
                                PDT["B{}".format(linha)] = Qtd
                            elif Tamanho in PDT['C2'].value:
                                PDT["C{}".format(linha)] = Qtd
                            elif Tamanho in PDT['D2'].value:
                                PDT["D{}".format(linha)] = Qtd
                            elif Tamanho in PDT['E2'].value:
                                PDT["E{}".format(linha)] = Qtd
                            elif Tamanho in PDT['F2'].value:
                                PDT["F{}".format(linha)] = Qtd

        #Estruturas BT primária
        for i in range (n_estruturas_BT_lvl1):
            classificação = Estruturas_BT_lvl1[i+1].split(" - ")
            Estrutura = classificação [0]            
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PDT['A']:
                linha = celulal.row
                Tipo = PDT['A{}'.format(linha)].value
                if linha <= 65:
                    time.sleep(0.0000001)
                elif linha == 100:
                    break
                else:
                    if Tipo == "ESTRUTURA DE BT":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PDT['B2'].value:
                                PDT["B{}".format(linha)] = Qtd
                            elif Tamanho in PDT['C2'].value:
                                PDT["C{}".format(linha)] = Qtd
                            elif Tamanho in PDT['D2'].value:
                                PDT["D{}".format(linha)] = Qtd
                            elif Tamanho in PDT['E2'].value:
                                PDT["E{}".format(linha)] = Qtd
                            elif Tamanho in PDT['F2'].value:
                                PDT["F{}".format(linha)] = Qtd

        #Equipamentos primária
        for i in range (n_equipamentos):
            classificação = Equipamentos[i+1].split(" - ")
            Tipo_equipamento = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PDT['A']:
                linha = celulal.row
                Tipo = PDT['A{}'.format(linha)].value
                if linha <= 100:
                    time.sleep(0.0000001)
                elif linha == 120:
                    break
                else:
                    if Tipo == "EQUIP.":
                        time.sleep(0.001)
                    else:
                        if Tipo == Tipo_equipamento:
                            if Tamanho in PDT['B2'].value:
                                PDT["B{}".format(linha)] = Qtd
                            elif Tamanho in PDT['C2'].value:
                                PDT["C{}".format(linha)] = Qtd
                            elif Tamanho in PDT['D2'].value:
                                PDT["D{}".format(linha)] = Qtd
                            elif Tamanho in PDT['E2'].value:
                                PDT["E{}".format(linha)] = Qtd
                            elif Tamanho in PDT['F2'].value:
                                PDT["F{}".format(linha)] = Qtd

        #Estruturas MT Secundária
        for i in range (n_estruturas_MT_lvl2):
            classificação = Estruturas_MT_lvl2[i+2].split(" - ")
            Estrutura = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PDT['A']:
                linha = celulal.row
                Tipo = PDT['A{}'.format(linha)].value
                if linha <= 2:
                    time.sleep(0.0000001)
                elif linha == 65:
                    break
                else:
                    if Tipo == "ESTRUTURAS E OS POSTES ONDE ESTÃO INSTALADOS DO 1º NÍVEL" or Tipo == "ESTRUTURA DE MT":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PDT['G2'].value:
                                PDT["G{}".format(linha)] = Qtd
                            elif Tamanho in PDT['H2'].value:
                                PDT["H{}".format(linha)] = Qtd
                            elif Tamanho in PDT['I2'].value:
                                PDT["I{}".format(linha)] = Qtd
                            elif Tamanho in PDT['J2'].value:
                                PDT["J{}".format(linha)] = Qtd
                            elif Tamanho in PDT['K2'].value:
                                PDT["K{}".format(linha)] = Qtd

        #Estruturas BT Secundária
        for i in range (n_estruturas_BT_lvl2):
            classificação = Estruturas_BT_lvl2[i+1].split(" - ")
            Estrutura = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PDT['A']:
                linha = celulal.row
                Tipo = PDT['A{}'.format(linha)].value
                if linha <= 65:
                    time.sleep(0.0000001)
                elif linha == 100:
                    break
                else:
                    if Tipo == "ESTRUTURA DE BT":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PDT['G2'].value:
                                PDT["G{}".format(linha)] = Qtd
                            elif Tamanho in PDT['H2'].value:
                                PDT["H{}".format(linha)] = Qtd
                            elif Tamanho in PDT['I2'].value:
                                PDT["I{}".format(linha)] = Qtd
                            elif Tamanho in PDT['J2'].value:
                                PDT["J{}".format(linha)] = Qtd
                            elif Tamanho in PDT['K2'].value:
                                PDT["K{}".format(linha)] = Qtd
    DT()

    def RC():

        #Dividir níveis de Estrutura
        First_lvl = splitRC[0]
        Second_lvl = splitRC[1]

        #Dividir MT e BT de cada Nível
        MT_BT_lvl1 = First_lvl.split("\n\n")
        MT_BT_lvl2 = Second_lvl.split("\n\n")
        
        #MT e BT de cada Nível
        Estruturas_MT_lvl1 = MT_BT_lvl1[0].split("\n")
        Estruturas_BT_lvl1 = MT_BT_lvl1[1].split("\n")
        Equipamentos = MT_BT_lvl1[2].split("\n")
        Estruturas_MT_lvl2 = MT_BT_lvl2[0].split("\n")
        Estruturas_BT_lvl2 = MT_BT_lvl2[1].split("\n")

        #Número de estruturas
        n_estruturas_MT_lvl1 = int(len(Estruturas_MT_lvl1)-2)
        n_estruturas_BT_lvl1 = int(len(Estruturas_BT_lvl1)-1)
        n_equipamentos = int(len(Equipamentos)-1)
        n_estruturas_MT_lvl2 = int(len(Estruturas_MT_lvl2)-2)
        n_estruturas_BT_lvl2 = int(len(Estruturas_BT_lvl2)-1)

        #Estruturas MT primária
        for i in range (n_estruturas_MT_lvl1):
            classificação = Estruturas_MT_lvl1[i+2].split(" - ")
            Estrutura = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PRC['A']:
                linha = celulal.row
                Tipo = PRC['A{}'.format(linha)].value
                if linha == 1:
                    time.sleep(0.0000001)
                elif linha == 53:
                    break
                else:
                    if Tipo == "ESTRUTURA":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PRC['B1'].value:
                                PRC["B{}".format(linha)] = Qtd
                            elif Tamanho in PRC['C1'].value:
                                PRC["C{}".format(linha)] = Qtd
                            elif Tamanho in PRC['D1'].value:
                                PRC["D{}".format(linha)] = Qtd
                            elif Tamanho in PRC['E1'].value:
                                PRC["E{}".format(linha)] = Qtd
                            elif Tamanho in PRC['F1'].value:
                                PRC["F{}".format(linha)] = Qtd

        #Estruturas BT primária
        for i in range (n_estruturas_BT_lvl1):
            classificação = Estruturas_BT_lvl1[i+1].split(" - ")
            Estrutura = classificação [0]            
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PRC['A']:
                linha = celulal.row
                Tipo = PRC['A{}'.format(linha)].value
                if linha <= 52:
                    time.sleep(0.0000001)
                elif linha == 77:
                    break
                else:
                    if Tipo == "ESTRUTURA DE BT":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PRC['B1'].value:
                                PRC["B{}".format(linha)] = Qtd
                            elif Tamanho in PRC['C1'].value:
                                PRC["C{}".format(linha)] = Qtd
                            elif Tamanho in PRC['D1'].value:
                                PRC["D{}".format(linha)] = Qtd
                            elif Tamanho in PRC['E1'].value:
                                PRC["E{}".format(linha)] = Qtd
                            elif Tamanho in PRC['F1'].value:
                                PRC["F{}".format(linha)] = Qtd

        #Equipamentos primária
        for i in range (n_equipamentos):
            classificação = Equipamentos[i+1].split(" - ")
            Tipo_equipamento = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PRC['A']:
                linha = celulal.row
                Tipo = PRC['A{}'.format(linha)].value
                if linha <= 76:
                    time.sleep(0.0000001)
                elif linha == 84:
                    break
                else:
                    if Tipo == "EQUIP.":
                        time.sleep(0.001)
                    else:
                        if Tipo == Tipo_equipamento:
                            if Tamanho in PRC['B1'].value:
                                PRC["B{}".format(linha)] = Qtd
                            elif Tamanho in PRC['C1'].value:
                                PRC["C{}".format(linha)] = Qtd
                            elif Tamanho in PRC['D1'].value:
                                PRC["D{}".format(linha)] = Qtd
                            elif Tamanho in PRC['E1'].value:
                                PRC["E{}".format(linha)] = Qtd
                            elif Tamanho in PRC['F1'].value:
                                PRC["F{}".format(linha)] = Qtd

        #Estruturas MT Secundária
        for i in range (n_estruturas_MT_lvl2):
            classificação = Estruturas_MT_lvl2[i+2].split(" - ")
            Estrutura = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PRC['A']:
                linha = celulal.row
                Tipo = PRC['A{}'.format(linha)].value
                if linha <= 1:
                    time.sleep(0.0000001)
                elif linha == 53:
                    break
                else:
                    if Tipo == "ESTRUTURAS E OS POSTES ONDE ESTÃO INSTALADOS DO 1º NÍVEL" or Tipo == "ESTRUTURA DE MT":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PRC['G1'].value:
                                PRC["G{}".format(linha)] = Qtd
                            elif Tamanho in PRC['H1'].value:
                                PRC["H{}".format(linha)] = Qtd
                            elif Tamanho in PRC['I1'].value:
                                PRC["I{}".format(linha)] = Qtd
                            elif Tamanho in PRC['J1'].value:
                                PRC["J{}".format(linha)] = Qtd

        #Estruturas BT Secundária
        for i in range (n_estruturas_BT_lvl2):
            classificação = Estruturas_BT_lvl2[i+1].split(" - ")
            Estrutura = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PRC['A']:
                linha = celulal.row
                Tipo = PRC['A{}'.format(linha)].value
                if linha <= 52:
                    time.sleep(0.0000001)
                elif linha == 77:
                    break
                else:
                    if Tipo == "ESTRUTURA DE BT":
                        time.sleep(0.001)
                    else:
                        if Tipo == Estrutura:
                            if Tamanho in PRC['G1'].value:
                                PRC["G{}".format(linha)] = Qtd
                            elif Tamanho in PRC['H1'].value:
                                PRC["H{}".format(linha)] = Qtd
                            elif Tamanho in PRC['I1'].value:
                                PRC["I{}".format(linha)] = Qtd
                            elif Tamanho in PRC['J1'].value:
                                PRC["J{}".format(linha)] = Qtd

        #Equipamentos Secundária
        for i in range (n_equipamentos):
            classificação = Equipamentos[i+1].split(" - ")
            Tipo_equipamento = classificação [0]
            Tamanho = classificação [1]
            Qtd = int(classificação [2])


            #Escrever na Base
            for celulal in PRC['A']:
                linha = celulal.row
                Tipo = PRC['A{}'.format(linha)].value
                if linha <= 76:
                    time.sleep(0.0000001)
                elif linha == 84:
                    break
                else:
                    if Tipo == "EQUIP.":
                        time.sleep(0.001)
                    else:
                        if Tipo == Tipo_equipamento:
                            if Tamanho in PRC['G1'].value:
                                PRC["G{}".format(linha)] = Qtd
                            elif Tamanho in PRC['H1'].value:
                                PRC["H{}".format(linha)] = Qtd
                            elif Tamanho in PRC['I1'].value:
                                PRC["I{}".format(linha)] = Qtd
                            elif Tamanho in PRC['J1'].value:
                                PRC["J{}".format(linha)] = Qtd
    RC()

    base.save("C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra))
        
    excel_app = xlwings.App(visible = False)
    excel_book = excel_app.books.open("C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra))
    excel_book.save()
    excel_book.close()
    excel_app.quit()

    def Escrita():

        #Abrir Planilha Base e escrever no Arquivo   
        resultado_base = load_workbook ("C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra), data_only=True)
        Resultado = resultado_base["RESUMO DE MATERIAIS"]

        #Executar Escrita da Saída
        caminho_saida.write("Resumo dos Materiais:\n\n\n")
        i=0

        #Escrita Primeira Coluna
        for celula in Resultado['A']:
            linha = celula.row
            Descrição = Resultado['A{}'.format(linha)].value
            Qtd = Resultado['B{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 26:
                break
            else:
                caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                i+=1

        for celula in Resultado['A']:
            linha = celula.row
            Descrição = Resultado['A{}'.format(linha)].value
            Qtd = Resultado['B{}'.format(linha)].value
            if linha <= 27:
                time.sleep(0.0000001)
            elif linha == 34:
                break
            else:
                if linha == 29:
                    caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                    i+=1
                else:
                    caminho_saida.write("{}. {}: 'Entrada Manual'\n".format(i, Descrição))
                    i+=1

        #Escrita Segunda Coluna
        for celula in Resultado['C']:
            linha = celula.row
            Descrição = Resultado['C{}'.format(linha)].value
            Qtd = Resultado['D{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 37:
                break
            else:
                caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                i+=1

        #Escrita Terceira Coluna
        for celula in Resultado['E']:
            linha = celula.row
            Descrição = Resultado['E{}'.format(linha)].value
            Qtd = Resultado['F{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 37:
                break
            else:
                caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                i+=1   

        #Escrita Quarta Coluna
        for celula in Resultado['G']:
            linha = celula.row
            Descrição = Resultado['G{}'.format(linha)].value
            Qtd = Resultado['H{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 32:
                break
            else:
                if linha == 2:
                    caminho_saida.write("\n{}\n\n".format(Descrição))
                elif linha == 20:
                    caminho_saida.write("\n{}...{}\n\n".format(Descrição,Qtd))
                else:
                    caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                    i+=1             
    Escrita()

    caminho_saida.close()#Salvar informações

    '''#Apagar planilha resultado
    caminho = "C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra)
    caminho=os.path.realpath(caminho)
    os.remove(caminho)'''
    
main()