from openpyxl import Workbook, load_workbook
import time
import xlwings
import os

def main():

    usuario = str(input("Qual o seu usuário?\n"))
    n_Obra = str(input("\nQual o número da obra?\n"))
    
    #Base de dados
    base = load_workbook("Base de Dados - Obras.xlsx")
    entrada = load_workbook("C:\\Users\\{}\\Documents\\Medição\\{}\\Entrada.xlsx".format(usuario, n_Obra))
    PDT = base["PARAFUSOS DT"]
    PRC = base["PARAFUSOS RC"]
    DTE = entrada["DT"]
    RCE = entrada["RC"]
    
    caminho_saida =open("C:\\Users\\{}\\Documents\\Medição\\{}\\Saida.txt".format(usuario, n_Obra), 'w') # Saída do resultado
    
    def DT():

        colunas = ['A','B','C','D','E','F','G','H','I','J','K']
        a = 1

        for linha in DTE['A']:
            a = a + 1
            
            #Cabeçalho
            if a <= 2:
                time.sleep(0.0001)

            #Estruturas de MT
            elif a >= 3 and a < 65:
                for coluna in range(1,12):
                    if coluna == 1:
                        time.sleep(0.0001)
                    elif coluna >=2 and coluna < 7:
                        Qtd1 = DTE.cell(row = a, column = coluna).value
                        PDT['{}{}'.format(colunas[coluna-1],a)]= Qtd1
                    elif coluna >=7:
                        Qtd2 = DTE.cell(row = a, column = coluna).value
                        PDT['{}{}'.format(colunas[coluna-1],a)]= Qtd2
            
            #Estruturas de BT
            elif a >= 66 and a < 100:
                for coluna in range(1,12):
                    if coluna == 1:
                        time.sleep(0.0001)
                    elif coluna >=2 and coluna < 7:
                        Qtd1 = DTE.cell(row = a, column = coluna).value
                        PDT['{}{}'.format(colunas[coluna-1],a)]= Qtd1
                    elif coluna >=7:
                        Qtd2 = DTE.cell(row = a, column = coluna).value
                        PDT['{}{}'.format(colunas[coluna-1],a)]= Qtd2

            #Equipamentos da Estrutura
            elif a >= 101 and a < 120:
                for coluna in range(1,7):
                    if coluna == 1:
                        time.sleep(0.0001)
                    elif coluna >=2 and coluna < 7:
                        Qtd1 = DTE.cell(row = a, column = coluna).value
                        PDT['{}{}'.format(colunas[coluna-1],a)]= Qtd1
                    elif coluna >=7:
                        Qtd2 = DTE.cell(row = a, column = coluna).value
                        PDT['{}{}'.format(colunas[coluna-1],a)]= Qtd2
   
    DT()

    def RC():

        colunas = ['A','B','C','D','E','F','G','H','I','J']
        a = 1

        for linha in DTE['A']:
            a = a + 1

            #Cabeçalho
            if a <= 2:
                time.sleep(0.0001)

            #Estruturas de MT
            elif a >= 3 and a < 54:
                for coluna in range(1,11):
                    if coluna == 1:
                        time.sleep(0.0001)
                    elif coluna >=2 and coluna < 7:
                        Qtd1 = RCE.cell(row = a, column = coluna).value
                        PRC['{}{}'.format(colunas[coluna-1],a-1)]= Qtd1
                    elif coluna >=7:
                        Qtd2 = RCE.cell(row = a, column = coluna).value
                        PRC['{}{}'.format(colunas[coluna-1],a-1)]= Qtd2
            
            #Estruturas de BT
            elif a >= 55 and a < 79:
                for coluna in range(1,11):
                    if coluna == 1:
                        time.sleep(0.0001)
                    elif coluna >=2 and coluna < 7:
                        Qtd1 = RCE.cell(row = a, column = coluna).value
                        PRC['{}{}'.format(colunas[coluna-1],a-2)]= Qtd1
                    elif coluna >=7:
                        Qtd2 = RCE.cell(row = a, column = coluna).value
                        PRC['{}{}'.format(colunas[coluna-1],a-2)]= Qtd2

            #Equipamentos da Estrutura
            elif a >= 80 and a < 87:
                for coluna in range(1,11):
                    if coluna == 1:
                        time.sleep(0.0001)
                    elif coluna >=2 and coluna < 7:
                        Qtd1 = RCE.cell(row = a, column = coluna).value
                        PRC['{}{}'.format(colunas[coluna-1],a-3)]= Qtd1
                    elif coluna >=7:
                        Qtd2 = RCE.cell(row = a, column = coluna).value
                        PRC['{}{}'.format(colunas[coluna-1],a-3)]= Qtd2
   
    RC()

    def salvar_tabela():

        base.save("C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra))

        excel_app = xlwings.App(visible = False)
        excel_book = excel_app.books.open("C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra))
        excel_book.save()
        excel_book.close()
        excel_app.quit()

    salvar_tabela()

    def Escrita():

        #Abrir Planilha Base e escrever no Arquivo   
        resultado_base = load_workbook ("C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra), data_only=True)
        Resultado = resultado_base["RESUMO DE MATERIAIS"]

        #Executar Escrita da Saída
        caminho_saida.write("Resumo dos Materiais:\n\n\n")
        i=0

        #Escrita Primeira Coluna
        for celula in Resultado['A']:
            linha = celula.row
            Descrição = Resultado['A{}'.format(linha)].value
            Qtd = Resultado['B{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 26:
                break
            else:
                caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                i+=1

        for celula in Resultado['A']:
            linha = celula.row
            Descrição = Resultado['A{}'.format(linha)].value
            Qtd = Resultado['B{}'.format(linha)].value
            if linha <= 27:
                time.sleep(0.0000001)
            elif linha == 34:
                break
            else:
                if linha == 29:
                    caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                    i+=1
                else:
                    caminho_saida.write("{}. {}: 'Entrada Manual'\n".format(i, Descrição))
                    i+=1

        #Escrita Segunda Coluna
        for celula in Resultado['C']:
            linha = celula.row
            Descrição = Resultado['C{}'.format(linha)].value
            Qtd = Resultado['D{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 37:
                break
            else:
                caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                i+=1

        #Escrita Terceira Coluna
        for celula in Resultado['E']:
            linha = celula.row
            Descrição = Resultado['E{}'.format(linha)].value
            Qtd = Resultado['F{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 37:
                break
            else:
                caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                i+=1   

        #Escrita Quarta Coluna
        for celula in Resultado['G']:
            linha = celula.row
            Descrição = Resultado['G{}'.format(linha)].value
            Qtd = Resultado['H{}'.format(linha)].value
            if linha == 1:
                time.sleep(0.0000001)
            elif linha == 32:
                break
            else:
                if linha == 2:
                    caminho_saida.write("\n{}\n\n".format(Descrição))
                elif linha == 20:
                    caminho_saida.write("\n{}...{}\n\n".format(Descrição,Qtd))
                else:
                    caminho_saida.write("{}. {}: {}\n".format(i, Descrição, Qtd))
                    i+=1             
    
    Escrita()

    def finalizar():

        caminho_saida.close()#Salvar informações

        #Apagar planilha resultado
        caminho = "C:\\Users\\{}\\Documents\\Medição\\{}\\Mão de Obra - Almoxarifado - {}.xlsx".format(usuario,n_Obra,n_Obra)
        caminho=os.path.realpath(caminho)
        os.remove(caminho)

    finalizar()

main()